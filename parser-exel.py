import openpyxl as op
import time
import CONSTANTS



print('\nНапоминание! Все обрабатываемые файлы должны находиться в подпапке /tables/\n\
Все загружаемые файлы должны иметь расширение .xlsx !\n\
В противном случае высока вероятность ошибок!\n')

file_ozon = input('Укажите название опорного файла OZON и нажмите Enter: ')
file_nhl = input('Укажите название файла NHL и нажмите Enter: ')
file_khl = input('Укажите название файла KHL и нажмите Enter: ')
file_russia = input('Укажите название файла RUSSIA и нажмите Enter: ')
file_football = input('Укажите название файла FOOTBALL и нажмите Enter: ')
file_customer_order = input('Укажите название файла CUSTOMER_ORDER и нажмите Enter: ')



# //////////////////////////////////////////// PIVOT TABLE //////////////////////////////////////////////

print(f'Открываем файл {file_ozon}...')
file_ozon_pars = f'tables\\{file_ozon}'
wb_ozon = op.load_workbook(file_ozon_pars, data_only=True)
sheet_ozon = wb_ozon.active
max_rows_ozon = sheet_ozon.max_row
ozon_keys_list = []
ozon_dict = {}


print(f'Открываем {file_ozon}, импортируем таблицы...')
for row in sheet_ozon.iter_rows(min_row=2, max_row=max_rows_ozon, values_only=True):
    sku_ozon = str(row[0]).split('-')
    if len(sku_ozon) > 1:
        sku_ozon_f = '-'.join([sku_ozon[0], sku_ozon[-1]])
        ozon_dict[sku_ozon_f] = []
        ozon_keys_list.append(sku_ozon_f)
    else:
        sku_ozon_f = sku_ozon[0]
        ozon_dict[sku_ozon_f] = []
        ozon_keys_list.append(sku_ozon_f)

    extend_list = [row[3], 0, int(row[4])]
    ozon_dict[sku_ozon_f].extend(extend_list)



print('Создаем новый лист "result"...')
wb_ozon.create_sheet('result')
result_sheet = wb_ozon['result']
result_sheet.append(CONSTANTS.COLUMN_NAME)


print('Создаем опорную таблицу...')
rw = 2
cn = 1

for i in range(0, len(ozon_keys_list)):
    key = ozon_keys_list[i]
    for value in ozon_dict[key]:
        result_sheet.cell(row=rw, column=1).value = str(key)
        result_sheet.cell(row=rw, column=cn+1).value = int(value) if isinstance(value, int) else str(value)
        cn += 1
    rw += 1
    cn = 1

result_max_row = result_sheet.max_row


# //////////////////////////////////////////// NHL //////////////////////////////////////////////

print(f'Открываем {file_nhl}, импортируем таблицы...')
file_nhl_pars = f'tables\\{file_nhl}'
wb_nhl = op.load_workbook(file_nhl_pars, data_only=True)
sheet_nhl = wb_nhl.active
max_rows_nhl = sheet_nhl.max_row
nhl_dict = {}

wb_ozon.create_sheet('NHL')
nhl_sheet = wb_ozon['NHL']
nhl_sheet.append(CONSTANTS.COLUMN_NAME_ADDITIONAL_SHEET)


print(f'Извлечение значений из таблиц {file_nhl}...')
for row_nhl in sheet_nhl.iter_rows(min_row=8, max_row=max_rows_nhl, values_only=True):
    row_nhl_pk = row_nhl[14] if row_nhl[14] != None else 'None'
    while row_nhl_pk.startswith('0'):
        row_nhl_pk = row_nhl_pk[1:]
    if row_nhl_pk == 'None':
        continue
    elif str(row_nhl_pk).split(' ')[1] in CONSTANTS.SKU_NOSIZE_TYPE:
        sku_nhl = str(row_nhl_pk).split(' ')
        sku_nhl_f = sku_nhl[0]
        print(sku_nhl_f + '- Done')
        nhl_dict[sku_nhl_f] = row_nhl[16]
    else:
        sku_nhl = str(row_nhl_pk).split(' ')
        sku_nhl_v = sku_nhl[-3]
        sku_nhl_v = sku_nhl_v.replace('-', '/') if '-' in sku_nhl_v else sku_nhl_v
        sku_nhl_f = '-'.join([sku_nhl[0], sku_nhl_v])
        print(sku_nhl_f + '- Done!')
        nhl_dict[sku_nhl_f] = row_nhl[16]
    nhl_add_sheet = [sku_nhl_f, row_nhl[14], row_nhl[16]]
    nhl_sheet.append(nhl_add_sheet)


print(f'Сопоставление значений таблиц {file_nhl}...')
i_nhl_add = 2
for row_chek_nhl in result_sheet.iter_rows(min_row=2, max_row=result_max_row, values_only=True):
    if row_chek_nhl[0] in nhl_dict:
        value = nhl_dict[row_chek_nhl[0]]
        result_sheet.cell(row=i_nhl_add, column=5).value = int(value)
        i_nhl_add +=1
    else:
        result_sheet.cell(row=i_nhl_add, column=5).value = 0
        i_nhl_add +=1
    
# //////////////////////////////////////////// KHL //////////////////////////////////////////////

print(f'Открываем {file_khl}, импортируем таблицы...')
file_khl_pars = f'tables\\{file_khl}'
wb_khl = op.load_workbook(file_khl_pars, data_only=True)
sheet_khl = wb_khl.active
max_rows_khl = sheet_khl.max_row
khl_dict = {}

wb_ozon.create_sheet('KHL')
khl_sheet = wb_ozon['KHL']
khl_sheet.append(CONSTANTS.COLUMN_NAME_ADDITIONAL_SHEET)


print(f'Извлечение значений из таблиц {file_khl}...')
for row_khl in sheet_khl.iter_rows(min_row=8, max_row=max_rows_khl, values_only=True):
    row_khl_pk = row_khl[14] if row_khl[14] != None else 'None'
    while row_khl_pk.startswith('0'):
        row_khl_pk = row_khl_pk[1:]
    if row_khl_pk == 'None':
        continue
    elif str(row_khl_pk).split(' ')[1] in CONSTANTS.SKU_NOSIZE_TYPE:
        sku_khl = str(row_khl_pk).split(' ')
        sku_khl_f = sku_khl[0]
        print(sku_khl_f + '- Done!')
        khl_dict[sku_khl_f] = row_khl[16]
    else:
        sku_khl = str(row_khl_pk).split(' ')
        sku_khl_v = sku_khl[-3]
        sku_khl_v = sku_khl_v.replace('-', '/') if '-' in sku_khl_v else sku_khl_v
        sku_khl_f = '-'.join([sku_khl[0], sku_nhl_v])
        print(sku_khl_f + '- Done!')
        khl_dict[sku_khl_f] = row_khl[16]
    khl_add_sheet = [sku_khl_f, row_khl[14], row_khl[16]]
    khl_sheet.append(khl_add_sheet)


print(f'Сопоставление значений таблиц {file_khl}...')
i_khl_add = 2
for row_chek_khl in result_sheet.iter_rows(min_row=2, max_row=result_max_row, values_only=True):
    if row_chek_khl[0] in khl_dict:
        value = khl_dict[row_chek_khl[0]]
        result_sheet.cell(row=i_khl_add, column=6).value = int(value)
        i_khl_add +=1
    else:
        result_sheet.cell(row=i_khl_add, column=6).value = 0
        i_khl_add +=1

# //////////////////////////////////////////// RUSSIA //////////////////////////////////////////////

print(f'Открываем {file_russia}, импортируем таблицы...')
file_russia_pars = f'tables\\{file_russia}'
wb_russia = op.load_workbook(file_russia_pars, data_only=True)
sheet_russia = wb_russia.active
max_rows_russia = sheet_russia.max_row
russia_dict = {}

wb_ozon.create_sheet('RUSSIA')
russia_sheet = wb_ozon['RUSSIA']
russia_sheet.append(CONSTANTS.COLUMN_NAME_ADDITIONAL_SHEET)


print(f'Извлечение значений из таблиц {file_russia}...')
for row_russia in sheet_russia.iter_rows(min_row=8, max_row=max_rows_russia, values_only=True):
    row_russia_pk = row_russia[14] if row_russia[14] != None else 'None'
    while row_russia_pk.startswith('0'):
        row_russia_pk = row_russia_pk[1:]
    if row_russia_pk == 'None':
        continue
    elif str(row_russia_pk).split(' ')[1] in CONSTANTS.SKU_NOSIZE_TYPE:
        sku_russia = str(row_russia_pk).split(' ')
        sku_russia_f = sku_russia[0]
        print(sku_russia_f + '- Done!')
        russia_dict[sku_russia_f] = row_russia[16]
    else:
        sku_russia = str(row_russia_pk).split(' ')
        sku_russia_v = sku_russia[-3]
        sku_russia_v = sku_russia_v.replace('-', '/') if '-' in sku_russia_v else sku_russia_v
        sku_russia_f = '-'.join([sku_russia[0], sku_russia_v])
        print(sku_russia_f + '- Done!')
        russia_dict[sku_russia_f] = row_russia[16]
    russia_add_sheet = [sku_russia_f, row_russia[14], row_russia[16]]
    russia_sheet.append(russia_add_sheet)


i_russia_add = 2
print(f'Сопоставление значений таблиц {file_russia}...')
for row_chek_russia in result_sheet.iter_rows(min_row=2, max_row=result_max_row, values_only=True):
    if row_chek_russia[0] in russia_dict:
        value = russia_dict[row_chek_russia[0]]
        result_sheet.cell(row=i_russia_add, column=7).value = int(value)
        i_russia_add +=1
    else:
        result_sheet.cell(row=i_russia_add, column=7).value = 0
        i_russia_add +=1


# //////////////////////////////////////////// FOOTBALL //////////////////////////////////////////////

print(f'Открываем {file_football}, импортируем таблицы...')
file_football_pars = f'tables\\{file_football}'
wb_football = op.load_workbook(file_football_pars, data_only=True)
sheet_football = wb_football.active
max_rows_football = sheet_football.max_row
football_dict = {}

wb_ozon.create_sheet('FOOTBALL')
football_sheet = wb_ozon['FOOTBALL']
football_sheet.append(CONSTANTS.COLUMN_NAME_ADDITIONAL_SHEET)


print(f'Извлечение значений таблиц {file_football}...')
for row_football in sheet_football.iter_rows(min_row=8, max_row=max_rows_football, values_only=True):
    row_football_pk = row_football[14] if row_football[14] != None else 'None'
    while row_football_pk.startswith('0'):
        row_football_pk = row_football_pk[1:]
    if row_football_pk == 'None':
        continue
    elif str(row_football_pk).split(' ')[1] in CONSTANTS.SKU_NOSIZE_TYPE:
        sku_football = str(row_football_pk).split(' ')
        sku_football_f = sku_football[0]
        print(sku_football_f + '- Done!')
        football_dict[sku_football_f] = row_football[16]
    else:
        sku_football = str(row_football_pk).split(' ')
        sku_football_v = sku_russia[-3]
        sku_football_v = sku_football_v.replace('-', '/') if '-' in sku_football_v else sku_football_v
        sku_football_f = '-'.join([sku_football[0], sku_football_v])
        print(sku_football_f + '- Done!')
        football_dict[sku_football_f] = row_football[16]
    football_add_sheet = [sku_football_f, row_football[14], row_football[16]]
    football_sheet.append(football_add_sheet)


i_football_add = 2
print(f'Сопоставление значений таблиц {file_football}...')
for row_chek_football in result_sheet.iter_rows(min_row=2, max_row=result_max_row, values_only=True):
    if row_chek_football[0] in football_dict:
        value = football_dict[row_chek_football[0]]
        result_sheet.cell(row=i_football_add, column=8).value = int(value)
        i_football_add +=1
    else:
        result_sheet.cell(row=i_football_add, column=8).value = 0
        i_football_add +=1


# //////////////////////////////////////////// CUSTOMER ORDER //////////////////////////////////////////////

print(f'Открываем {file_customer_order}, импортируем таблицы...')
file_customer_order_pars = f'tables\\{file_customer_order}'
wb_customer = op.load_workbook(file_customer_order_pars, data_only=True)
sheet_customer = wb_customer.active
max_rows_customer = sheet_customer.max_row
customer_dict = {}

wb_ozon.create_sheet('CUSTOMER')
customer_sheet = wb_ozon['CUSTOMER']
customer_sheet.append(CONSTANTS.COLUMN_NAME_ADDITIONAL_SHEET)


for row_customer in sheet_customer.iter_rows(min_row=1, max_row=max_rows_customer, values_only=True):
    if isinstance(row_customer[1], int):
        row_cuctomer_pk = row_customer[17]
        while row_cuctomer_pk.startswith('0'):
            row_cuctomer_pk = row_cuctomer_pk[1:]
        if str(row_cuctomer_pk).split(' ')[1] in CONSTANTS.SKU_NOSIZE_TYPE:
            sku_customer = str(row_cuctomer_pk).split(' ')
            sku_customer_f = sku_customer[0]
            print(sku_customer_f + '- Done!')
            customer_dict[sku_customer_f] = row_customer[57]
        else:
            sku_customer = str(row_cuctomer_pk).split(' ')
            sku_customer_v = sku_customer[-3]
            sku_customer_v = sku_customer_v.replace('-', '/') if '-' in sku_customer_v else sku_customer_v
            sku_customer_f = '-'.join([sku_customer[0], sku_customer_v])
            print(sku_customer_f + '- Done!')
            customer_dict[sku_customer_f] = row_customer[57]
        customer_add_sheet = [sku_customer_f, row_customer[17], row_customer[57]]
        customer_sheet.append(customer_add_sheet)


i_customer_add = 2
print(f'Сопоставление значений таблиц {file_customer_order}...')
for row_chek_customer in result_sheet.iter_rows(min_row=2, max_row=result_max_row, values_only=True):
    if row_chek_customer[0] in customer_dict:
        value = customer_dict[row_chek_customer[0]]
        result_sheet.cell(row=i_customer_add, column=3).value = int(value)
        i_customer_add +=1
    else:
        result_sheet.cell(row=i_customer_add, column=3).value = 0
        i_customer_add +=1


# //////////////////////////////////////////// SUM COLUMN //////////////////////////////////////////////

print('Расчет столбца "Итог" ')
i_sum = 2
for sum_row in result_sheet.iter_rows(min_row=2, max_row=result_max_row, values_only=True):
    final_sum = sum([int(sum_row[4]), int(sum_row[5]), int(sum_row[6]), int(sum_row[7])])
    result_sheet.cell(row=i_sum, column=9).value = final_sum
    i_sum += 1



wb_ozon.save('new_file.xlsx')
print('Сохранение... ')
time.sleep(3)
final_value = input('Можете закрыть программу.')
